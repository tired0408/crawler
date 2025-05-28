"""
爱医助医的网站数据分析脚本1
"""
import os
import time
import shutil
import datetime
import openpyxl
import pandas as pd
from dateutil.relativedelta import relativedelta
from utils import init_chrome, read_multi_column
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC



def crawler_data_from_web(chrome_path, chromedriver_path, download_path, acount, password, search_info):
    driver = init_chrome(chromedriver_path, download_path, chrome_path=chrome_path, is_proxy=False)
    pattern = (By.XPATH, "//button[contains(text(), '登录')]")
    print("打开网页")
    driver.get(r"http://cbs.aiyizhuyi.com/ayzy/cbs/yearly_data/monthsum")
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable(pattern))
    print("输入账号密码")
    ele = driver.find_element(By.NAME, "login")
    ele.clear()
    ele.send_keys(acount)
    ele = driver.find_element(By.NAME, "password")
    ele.clear()
    ele.send_keys(password)
    print("登录系统")
    ele = driver.find_element(*pattern)
    ele.click()
    print("点击《数据报告》模块")
    pattern = (By.XPATH, "//div[contains(@id, 'layout-canvas')]//ul[contains(@class, 'mainmenu-nav')]//span[contains(text(), '数据报告')]")
    ele = WebDriverWait(driver, 10).until(EC.element_to_be_clickable(pattern))
    ele.click()
    pattern = (By.XPATH, "//span[contains(text(), '设备耗材商务系统展示')]")
    ele = WebDriverWait(driver, 10).until(EC.presence_of_element_located(pattern))
    print("滚动到《设备耗材商务系统展示》选择标签")
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", ele)
    ele = WebDriverWait(driver, 10).until(EC.element_to_be_clickable(pattern))
    print("点击《设备耗材商务系统展示》")
    ele.click()
    for year, month, region in search_info:
        print(f"查看当前是否为目标年份:{year}")
        ele = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "select2-year-container")))
        if year not in ele.text:
            print("年份有误,选择年份")
            ele.click()
            ele = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, f"//ul[@id='select2-year-results']//li[contains(text(), f'{year}')]")))
            ele.click()
            WebDriverWait(driver, 10).until(EC.text_to_be_present_in_element((By.ID, "select2-month-container"), f"{year}年"))
        print(f"查看当前是否为目标月份:{month}")
        ele = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "select2-month-container")))
        if month not in ele.text:
            print("月份有误,选择月份")
            ele.click()
            pattern = (By.XPATH, f"//ul[@id='select2-month-results']//li[contains(text(), '{int(month)}')]")
            ele = WebDriverWait(driver, 10).until(EC.presence_of_element_located(pattern))
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", ele)
            ele = WebDriverWait(driver, 10).until(EC.element_to_be_clickable(pattern))
            ele.click()
            WebDriverWait(driver, 10).until(EC.text_to_be_present_in_element((By.ID, "select2-month-container"), f"{int(month)}月"))
        print(f"查看当前区域是否为目标区域:{region}")
        ele = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "select2-region-container")))
        if region not in ele.text:
            print("区域有误,选择区域")
            ele.click()
            pattern = (By.XPATH, f"//ul[@id='select2-region-results']//li[contains(text(), f'{region}')]")
            ele = WebDriverWait(driver, 10).until(EC.presence_of_element_located(pattern))
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", ele)
            ele = WebDriverWait(driver, 10).until(EC.element_to_be_clickable(pattern))
            ele.click()
            WebDriverWait(driver, 10).until(EC.text_to_be_present_in_element((By.ID, "select2-region-container"), f"{region}"))
        print("导出EXCEL到目标文件夹")
        ele = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "download1")))
        ele.click()
        st = time.time()
        print("等待文件下载完成")
        name = f"{year}-{month}_{region}_全部设备耗材汇总.xlsx"
        file_path = os.path.join(download_path, name)
        while time.time() - st < 300:
            if os.path.exists(file_path):
                break
            time.sleep(5)
        print(f"{name}, 已下载完成")
    print("关闭浏览器")
    driver.quit()


def main(acount, password, date_range:str, regions:str):
    regions = regions.split(",")
    search_info = []
    start_date, end_date = date_range.split("-")
    start_date = datetime.datetime.strptime(start_date, "%Y.%m")
    end_date = datetime.datetime.strptime(end_date, "%Y.%m")
    while start_date <= end_date:
        for region in regions:
            search_info.append([start_date.strftime("%Y"), start_date.strftime("%m"), region])
        start_date += relativedelta(months=1)
    path = os.path.dirname(__file__)
    download_path = os.path.join(path, "file")
    chromedriver_path = os.path.join(path, r"..\chromedriver_mac_arm64_114\chromedriver.exe")
    chrome_path = os.path.join(path, r"..\chromedriver_mac_arm64_114\chrome114\App\Chrome-bin\chrome.exe")
    output_path = os.path.join(path, "设备耗材汇总.xlsx")
    if not os.path.exists(output_path):
        raise Exception("《设备耗材汇总》模板文件不存在")
    print("从网站下载数据")
    if os.path.exists(download_path):
        shutil.rmtree(download_path)
    os.makedirs(download_path)
    crawler_data_from_web(chrome_path, chromedriver_path, download_path, acount, password, search_info)
    print("整理下载后的EXCEL文件,并合并到《设备耗材汇总》表")
    output_df = read_multi_column(output_path)
    end_row = output_df.iloc[-1]
    if end_row.astype("str").str.contains("合计").any():
        output_df = output_df.iloc[:-1]
    print("将导出的数据合并到目标表中")
    for file in os.listdir(download_path):
        file_path = os.path.join(download_path, file)
        input_df = read_multi_column(file_path)
        input_df = input_df.iloc[:-1]
        output_df = pd.concat([output_df, input_df], ignore_index=True)
    print("删除重复数据")
    output_df = output_df.drop_duplicates()
    output_df = output_df.sort_values(by="时间", ascending=True, ignore_index=True)
    print("保存数据")
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    # 删除第3行开始的所有行（到最大行）
    ws.delete_rows(3, ws.max_row - 2)
    # 写入数据
    for row_i, row in output_df.iterrows():
        row = row.fillna("")
        for col_i, value in enumerate(row):
            ws.cell(row=row_i+3, column=col_i+1, value=value)
    wb.save(output_path)
    print("《设备耗材汇总》表已完成")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("-a", "--acount", type=str, help="爱医助医的账号")
    parser.add_argument("-p", "--password", type=str, help="爱医助医的密码")
    parser.add_argument("-d", "--date_range", type=str, help="时间范围, 格式: 2025.01-2025.04")
    parser.add_argument("-r", "--regions", type=str, help="区域, 格式: 其他区域,四川区域")
    opt = parser.parse_args()
    main(opt.acount, opt.password, opt.date_range, opt.regions)