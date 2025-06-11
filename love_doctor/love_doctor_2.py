"""
爱医助医的网站数据分析脚本2
"""
import os
import time
import shutil
import datetime
import xlsxwriter
import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook
from typing import Dict
from utils import read_multi_column, read_by_openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC



user_folder = os.path.dirname(__file__)
now_time = datetime.datetime.now()
print("读取《设备耗材汇总.xlsx》")
wb, ws, header = read_by_openpyxl(os.path.join(user_folder, "设备耗材汇总.xlsx"))
print("分析《设备耗材汇总.xlsx》")
ws_max_row = ws.max_row
output_dict: Dict[str, Dict] = {}
end_time = None
st = time.time()
for row in tqdm(ws.iter_rows(min_row=3, values_only=True), total=ws_max_row-2, desc="分析中"):
    row_dict = dict(zip(header, row))
    sum_data = row_dict["本区域流向"] + row_dict["跨区域流向"]
    key = f'{row_dict["归属区域"]}_{row_dict["省份"]}_{row_dict["客户"]}'
    if str(now_time.year) not in row_dict["时间"]:
        finish_year_work = 0 
        finish_area_work = 0 
    else:
        finish_year_work = row_dict["已完成系统工单数量"]
        finish_area_work = row_dict["区域系统工单数量"]
    if key not in output_dict:
        output_dict[key] = {row_dict["时间"]: sum_data, "本年累计已完成工单数": finish_year_work, "本年累计区域系统工单数": finish_area_work}
    else:
        output_dict[key][row_dict["时间"]] = sum_data
        output_dict[key]["本年累计已完成工单数"] += finish_year_work
        output_dict[key]["本年累计区域系统工单数"] += finish_area_work
    end_time = row_dict["时间"]
wb.close()
end_time = datetime.datetime.strptime(end_time, "%Y-%m")
print("读取《数据库-客户分类目录.xlsx》")
wb, ws, header = read_by_openpyxl(os.path.join(user_folder, "数据库-客户分类目录.xlsx"), header=1)
print("分析《数据库-客户分类目录.xlsx》")
ws_max_row = ws.max_row
name2classify = {}
for row in tqdm(ws.iter_rows(min_row=2, values_only=True), total=ws_max_row-1, desc="分析中"):
    row_dict = dict(zip(header, row))
    name2classify[row_dict["客户名称"]] = row_dict["终端分类"]
print("转换成《全部设备耗材流向跟踪表》所需的数据, 输出到新表格中")
wb = xlsxwriter.Workbook(os.path.join(user_folder, "全部设备耗材流向跟踪表.xlsx"))
ws = wb.add_worksheet()
# 创建格式对象
format_border = wb.add_format({
    'border': 1,  # 1=细边框
    'border_color': '#000000'  # 黑色边框
})

format_light_green = wb.add_format({
    'border': 1,
    'bg_color': '#C6EFCE',  # 淡绿色背景[1,5](@ref)
    'border_color': '#000000'
})

format_light_orange = wb.add_format({
    'border': 1,
    'bg_color': '#FFD966',  # 淡橙色背景（自定义）
    'border_color': '#000000'
})
# 构建标题
columns=["归属区域", "省份", "客户"]
for i in range(1, 13):
    columns.append(f"{now_time.year-1}-{i:02d}")
columns.append(f"{now_time.year-1}年合计")
for i in range(1, end_time.month+1):
    columns.append(f"{end_time.year}-{i:02d}")
columns.append(f"{end_time.year}年合计")
columns.append("终端状态")
columns.append("终端分类")
columns.append("本年累计已完成工单数")
columns.append("本年累计区域系统工单数")
columns.append("本年平均设备保养率")
for i, value in enumerate(columns):
    ws.write(0, i, value, format_border)
# 写入数据
row_i = 0
for key, value in tqdm(output_dict.items(), desc="转换中"):
    row_i += 1
    region, province, custom = key.split("_")
    ws.write(row_i, 0, region, format_border)
    ws.write(row_i, 1, province, format_border)
    ws.write(row_i, 2, custom, format_border)
    for i in range(3, 15):
        ws.write(row_i, i, value.get(columns[i], 0), format_border)
    last_year_sum = sum(value.get(columns[i], 0) for i in range(3, 15))
    ws.write(row_i, 15, last_year_sum, format_light_green)
    end_index = columns.index(f"{end_time.year}年合计")
    for i in range(16, end_index):
        ws.write(row_i, i, value.get(columns[i], 0), format_border)
    now_year_sum = sum(value.get(columns[i], 0) for i in range(16, end_index))
    ws.write(row_i, end_index, now_year_sum, format_light_green)
    # 计算终端状态
    if last_year_sum <= 0 and now_year_sum > 0:
        ws.write(row_i, columns.index("终端状态"), "新增", format_border)
    elif last_year_sum > 0 and now_year_sum> 0:
        ws.write(row_i, columns.index("终端状态"), "存量", format_border)
    elif last_year_sum > 0 and now_year_sum <= 0:
        ws.write(row_i, columns.index("终端状态"), "丢失", format_border)
    else:
        ws.write(row_i, columns.index("终端状态"), "未开发", format_border)
    ws.write(row_i, columns.index("终端分类"), name2classify.get(custom, ""), format_border)
    ws.write(row_i, columns.index("本年累计已完成工单数"), value["本年累计已完成工单数"], format_light_orange)
    ws.write(row_i, columns.index("本年累计区域系统工单数"), value["本年累计区域系统工单数"], format_light_orange)
    ratio = value["本年累计已完成工单数"]/value["本年累计区域系统工单数"]*100 if value["本年累计区域系统工单数"] > 0 else 0
    ws.write(row_i, columns.index("本年平均设备保养率"), f"{ratio:.2f}%", format_light_orange)
wb.close()
print("《全部设备耗材流向跟踪表》已构建完成")