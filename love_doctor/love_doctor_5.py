"""
爱医助医的相关数据的整理
数据报告-客户维护
"""
import os
import re
from typing import List, Tuple
from glob import glob
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string




def main():
    path = os.path.dirname(__file__)
    download_path = os.path.join(path, "custom_file")
    files = glob(os.path.join(download_path, "客户维护数据*.xlsx"))
    files = [path for path in files if "客户维护数据报告-汇总表.xlsx" not in path]
    wb = Workbook()
    ws = wb.active
    # 写入标题
    headers = ["省份", "终端客户"]
    for header_text in process_raw_data(files):
        if "合计" in header_text:
            headers.append(header_text)
        else:
            headers.extend(["是否跨省销售", "跨省销售信息", header_text])
    for col_i, header in enumerate(headers):
        ws.cell(row=1, column=col_i+1, value=header)
    # 填充数据
    row_index_sign = []
    for path in files:
        year_month = extract_ym_str(os.path.basename(path))
        col_i_time = headers.index(year_month) + 1
        col_i_across_province = col_i_time - 2
        col_i_across_detail = col_i_time - 1
        data_wb = load_workbook(path)
        data_ws = data_wb["客户维护数据报告"]
        province = data_ws.cell(row=2, column=1).value
        if province is None:
            print(f"文件 {path} 的省份信息缺失, 跳过")
            continue
        data_ws = data_wb["流向明细"]
        header_row = next(data_ws.iter_rows(min_row=1, max_row=1, values_only=True))
        # 写入所需数据
        purchase_col = header_row.index("耗材采购量")
        cross_province_col = header_row.index("是否跨省销售")
        cross_detail_col = header_row.index("跨省销售信息")
        for row in data_ws.iter_rows(min_row=2, values_only=True):
            # 获取行索引（从1开始）,如果没有则新增
            custom_name = row[0]
            sign = f"{province}&&&{custom_name}"
            if sign not in row_index_sign:
                row_index_sign.append(sign)
                row_i = len(row_index_sign) + 1
                ws.cell(row=row_i, column=1, value=province)
                ws.cell(row=row_i, column=2, value=custom_name)
            else:
                row_i = row_index_sign.index(sign) + 2
            # 填充数据
            ws.cell(row=row_i, column=col_i_time, value=row[purchase_col])
            ws.cell(row=row_i, column=col_i_across_province, value=row[cross_province_col])
            ws.cell(row=row_i, column=col_i_across_detail, value=row[cross_detail_col])
    # 填充背景色:合计行, 并计算合计数量
    color_indexs = [i for i, h in enumerate(headers) if "合计" in h]
    colors = ["FFFF99", "CCE5FF", "E2F0D9", "F2F2F2", "FFE6CC", "E5CCFF", "FFCCCC"][:len(color_indexs)]
    sum_start_index = 3
    for color, color_index in zip(colors, color_indexs):
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        for row in range(2, ws.max_row+1):
            start_letter = get_column_letter(sum_start_index)
            end_letter = get_column_letter(color_index)
            cell = ws.cell(row=row, column=color_index+1, value=f"=SUM({start_letter}{row}:{end_letter}{row})")
            cell.fill = fill
        sum_start_index = color_index + 1
    # 保存文件
    print("保存文件")
    wb.save(os.path.join(download_path, "客户维护数据报告-汇总表.xlsx"))
    
def extract_ym_str(text: str) -> str:
    """从字符串中提取 '2025年11月' 格式的年月字符串（第一个匹配）"""
    match = re.search(r'\d+年\d+月', text)
    if not match:
        raise ValueError(f"无法提取年月: {text}")
    return match.group(0)

def parse_ym(ym_str: str) -> Tuple[int, int]:
    """将 '2025年11月' 解析为 (年, 月) 元组"""
    match = re.match(r'(\d+)年(\d+)月', ym_str)
    if not match:
        raise ValueError(f"非法年月格式: {ym_str}")
    return int(match[1]), int(match[2])

def sort_and_add_year_summary(ym_list: List[str]) -> List[str]:
    """
    输入：年月字符串列表（如 ['2025年11月', '2025年12月', '2026年1月']）
    输出：排序后并插入年份合计的列表
    """
    # 按年月数值排序
    sorted_ym = sorted(ym_list, key=parse_ym)

    result = []
    current_year = None

    for ym in sorted_ym:
        year, _ = parse_ym(ym)
        if current_year is None:
            current_year = year
        elif year != current_year:
            # 插入上一年合计
            result.append(f"{current_year}年合计")
            current_year = year
        result.append(ym)

    # 最后一年的合计
    if current_year is not None:
        result.append(f"{current_year}年合计")

    return result

def process_raw_data(raw_strings: List[str]) -> List[str]:
    """
    完整流程：提取年月 -> 排序并添加合计
    """
    ym_list = list(set([extract_ym_str(s) for s in raw_strings]))
    return sort_and_add_year_summary(ym_list)



if __name__ == "__main__":
    main()