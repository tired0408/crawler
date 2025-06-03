"""
爬虫抓取的相关通用工具
"""
import pandas as pd
from openpyxl import load_workbook
from selenium.webdriver import Chrome
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options


def init_chrome(chromedriver_path, download_path, user_path=None, chrome_path=None, is_proxy=True):
    """初始化浏览器
    
    args:
        chromedriver_path: (str); 浏览器驱动的地址
        download_path: (str); 下载路径
        chrome_path: (str); 浏览器路径
        is_proxy: (bool); 是否使用代理
    """
    service = Service(chromedriver_path)
    options = Options()
    if user_path is not None:
        options.add_argument(f'user-data-dir={user_path}')  # 指定用户数据目录
    if chrome_path is not None:
        options.binary_location = chrome_path
    if is_proxy:
        options.add_argument('--proxy-server=127.0.0.1:8080')
        options.add_argument('ignore-certificate-errors')
    options.add_argument('--log-level=3')
    options.add_experimental_option('prefs', {
        "download.default_directory": download_path,  # 指定下载目录
    })
    driver = Chrome(service=service, options=options)
    return driver

def read_multi_column(path):
    """读取两级列名的EXCEL表格"""
    df = pd.read_excel(path, header=[0, 1], engine="openpyxl")
    df.columns = [c1 if c2.startswith("Unnamed") else c2 for c1, c2 in df.columns]
    return df

def read_by_openpyxl(path, header=2):
    """使用openpyxl读取excel, 两行列名的"""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if header == 2:
        header1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        header2 = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        header = [cell2 if cell2 is not None else cell1 for cell1, cell2 in zip(header1, header2)]
    else:
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return wb, ws, header